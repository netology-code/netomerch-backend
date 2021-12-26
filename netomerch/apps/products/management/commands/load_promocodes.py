from django.core.management.base import BaseCommand
from openpyxl import load_workbook
import difflib
from apps.products.models import Item
from apps.orders.models import Promocode


def find_match(sample, search_set):
    similarity = 0.5
    for item in search_set:
        matcher = difflib.SequenceMatcher(a=sample.lower(), b=item[0].lower())
        ratio = matcher.ratio()
        if ratio > similarity:
            similarity = ratio
            return item
    return False


def parse_xls(filename):
    item_sets = {(item.name, item) for item in Item.objects.all()}
    promo = dict()
    log = []

    wb = load_workbook(filename)
    sheet = wb[wb.sheetnames[0]]

    row = 2

    while row > 0:
        if not sheet.cell(column=1, row=row).value:
            break
        try:
            code = sheet.cell(column=1, row=row).value
            email = sheet.cell(column=2, row=row).value
            item_name = sheet.cell(column=3, row=row).value
            item = find_match(item_name, item_sets)
            if item is False:
                raise NameError

            item = item[1]
            if (code, email) not in promo:
                promo[(code, email)] = {
                    "code": code,
                    "email": email,
                    "item": item
                }
            else:
                raise IndexError
        except NameError as exc:
            log.append(f'Error in row {row}: {exc}')
        except IndexError as exc:
            log.append(f'Error in row {row}: {exc}')
        row += 1
    return promo, log

class Command(BaseCommand):
    help = 'Loads promo from xlsx file'

    def add_arguments(self, parser):
        parser.add_argument('filename', nargs=1, type=str)

    def handle(self, *args, **options):  # Noqa: C901
        promo, log = parse_xls(options['filename'][0])
        for item in promo.values():
            db_promo = Promocode.objects.create(code=item['code'],
                                                email=item['email'],
                                                item=item['item'])

            db_promo.save()

        return '\n'.join(log)
