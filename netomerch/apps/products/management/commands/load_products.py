import difflib

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from apps.products.models import Category, DictImageColor, ImageColorItem, Item, Size, Specialization
from apps.products.tasks import clear_deps, download_image


def find_match(sample, search_set):
    result = sample
    similarity = 0.5
    for item in search_set:
        matcher = difflib.SequenceMatcher(sample.lower(), item.lower())
        ratio = matcher.ratio()
        if ratio > similarity:
            similarity = ratio
            result = item
    return result


def strip_or_empty(string):
    return string.strip() if string is not None else ""


def parse_xlsx(filename):
    cat_set = {cat.name for cat in Category.objects.all()}
    color_set = {color.name for color in DictImageColor.objects.all()}
    spec_set = {spec.name for spec in Specialization.objects.all()}

    items = dict()
    log = []

    wb = load_workbook(filename)
    sheet = wb[wb.sheetnames[0]]

    row = 2

    while row > 0:
        if not sheet.cell(column=2, row=row).value:
            break
        try:
            default_color = sheet.cell(column=8, row=row).value is not None
            name = sheet.cell(column=2, row=row).value
            specialization = sheet.cell(column=4, row=row).value
            specialization = find_match(specialization, spec_set)
            spec_set.add(specialization)
            if (name, specialization) not in items:
                items[(name, specialization)] = {
                    "name": name,
                    "specialization": specialization,
                    "colors": [],
                }
            category = sheet.cell(column=3, row=row).value
            category = find_match(category, cat_set)
            cat_set.add(category)
            items[(name, specialization)]['category'] = category
            items[(name, specialization)]['sizes'] = [
                size.strip() for size in sheet.cell(column=5, row=row).value.split(',')
            ]
            color_tuple = [color.strip() for color in sheet.cell(column=6, row=row).value.split(',')]
            color_tuple[0] = find_match(color_tuple[0], color_set)
            color_set.add(color_tuple[0])
            items[(name, specialization)]['colors'].append({
                'color': {'name': color_tuple[0], 'code': color_tuple[1]},
                'images': [sheet.cell(column=j, row=row).value for j in range(9, 13)],
                'default': default_color
            })
            description = sheet.cell(column=7, row=row).value
            items[(name, specialization)]['description'] = strip_or_empty(description)
            items[(name, specialization)]['short_description'] = ""
            items[(name, specialization)]['is_hit'] = sheet.cell(column=13, row=row).value is not None
            price = sheet.cell(column=14, row=row).value
            items[(name, specialization)]['price'] = float(price) if price is not None else 0.0
        except Exception as exc:
            log.append(f'Error in row {row}: {exc}')
        row += 1
    return items, log


class Command(BaseCommand):
    help = 'Loads items from xlsx file'

    def add_arguments(self, parser):
        parser.add_argument('filename', nargs=1, type=str)

    def handle(self, *args, **options):  # Noqa: C901
        items, log = parse_xlsx(options['filename'][0])
        for item in items.values():
            category, _ = Category.objects.get_or_create(name=item['category'])

            specialization, _ = Specialization.objects.get_or_create(name=item['specialization'])

            db_item, _ = Item.objects.get_or_create(name=item['name'],
                                                    category=category,
                                                    specialization=specialization)

            db_item.description = item['description']
            db_item.short_description = item['short_description']

            db_item.price = item['price']
            db_item.is_hit = item['is_hit']
            db_item.save()

            clear_deps(db_item)

            for size in item['sizes']:
                db_size, _ = Size.objects.get_or_create(name=size.capitalize())
                db_item.size.add(db_size)

            for color in item['colors']:
                db_color, created = DictImageColor.objects.get_or_create(name=color['color']['name'].capitalize())
                if created:
                    db_color.color_code = color['color']['code']
                    db_color.save()
                main_image = True
                i = 0
                for image in color['images']:
                    tmp_image = f'no_image_{i}'
                    i += 1
                    db_image_color, _ = ImageColorItem.objects.get_or_create(
                        color=db_color,
                        item=db_item,
                        image=tmp_image)
                    db_image_color.is_main_image = main_image
                    db_image_color.is_main_color = color['default']
                    db_image_color.image = tmp_image
                    if color['default']:
                        color['default'] = False
                    db_image_color.save()
                    download_image.delay(db_image_color.id, image)
                    main_image = False
        return '\n'.join(log)
